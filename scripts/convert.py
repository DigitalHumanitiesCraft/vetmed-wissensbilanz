#!/usr/bin/env python3
"""
Wissensbilanz Konverter - Unified Excel zu JSON Konverter.

Konsolidiert alle Konvertierungslogik fuer Wissensbilanz-Daten.
Ersetzt: convert_excel_to_json.py, convert_3a1.py, convert_3a3.py

Verwendung:
    python scripts/convert.py                    # Alle Dateien konvertieren
    python scripts/convert.py --file "1-A-1"     # Spezifische Kennzahl
    python scripts/convert.py --analyze          # Struktur analysieren
    python scripts/convert.py --validate         # Output validieren

Autor: VetMed AI Initiative
Version: 2.0.0
"""

import os
import sys
import json
import re
import argparse
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Set

try:
    import openpyxl
except ImportError:
    print("[Converter] Fehler: openpyxl nicht installiert")
    print("[Converter] Installation: pip install openpyxl")
    sys.exit(1)


# ============================================================
# KONFIGURATION
# ============================================================

# Offizielle Uni-Codes aus der Wissensbilanz-Verordnung
VALID_UNI_CODES = {
    "UA": "Universitaet Wien",
    "UB": "Universitaet Graz",
    "UC": "Universitaet Innsbruck",
    "UD": "Universitaet Salzburg",
    "UE": "Technische Universitaet Wien",
    "UF": "Technische Universitaet Graz",
    "UG": "Montanuniversitaet Leoben",
    "UH": "Universitaet fuer Bodenkultur Wien",
    "UI": "Veterinaermedizinische Universitaet Wien",
    "UJ": "Wirtschaftsuniversitaet Wien",
    "UK": "Universitaet Linz",
    "UL": "Universitaet Klagenfurt",
    "UM": "Universitaet fuer Weiterbildung Krems",
    "UN": "Medizinische Universitaet Wien",
    "UO": "Medizinische Universitaet Graz",
    "UQ": "Medizinische Universitaet Innsbruck",
    "UR": "Akademie der bildenden Kuenste Wien",
    "US": "Universitaet fuer angewandte Kunst Wien",
    "UT": "Universitaet fuer Musik und darstellende Kunst Wien",
    "UU": "Universitaet Mozarteum Salzburg",
    "UV": "Universitaet fuer Musik und darstellende Kunst Graz",
    "UW": "Universitaet fuer kuenstlerische und industrielle Gestaltung Linz",
}

# Gueltiger Jahresbereich
VALID_YEARS = range(2019, 2030)

# Mapping Buchstabe -> Uni-Code (fuer 3-A-3)
LETTER_TO_UNI_CODE = {char: f"U{char}" for char in "ABCDEFGHIJKLMNOQRSTUVW"}

# Dateiname -> Kennzahl-Code Mapping
FILE_TO_KENNZAHL = {
    "1-A-1 Personal - Koepfe.xlsx": "1-A-1",
    "1-A-1 Personal - VZAe.xlsx": "1-A-1-VZA",
    "1-A-2 Berufungen an die Universitaet.xlsx": "1-A-2",
    "1-A-3 Frauenquote in Kollegialorganen.xlsx": "1-A-3",
    "1-A-4 Gender pay gap.xlsx": "1-A-4",
    "1-A-5 Repraesentanz von Frauen in Berufungsverfahren.xlsx": "1-A-5",
    "2-A-1 ProfessorInnen und Aequivalente.xlsx": "2-A-1",
    "2-A-2 Eingerichtete Studien.xlsx": "2-A-2",
    "2-A-3 Studienabschlussquote.xlsx": "2-A-3",
    "2-A-4 Besondere Zulassungsbedingungen.xlsx": "2-A-4",
    "2-A-5 Anzahl Studierenden.xlsx": "2-A-5",
    "2-A-6 Anzahl Pruefungsaktive.xlsx": "2-A-6",
    "2-A-7 Anzahl belegte ordentliche Studien.xlsx": "2-A-7",
    "2-A-8 Ordentliche Studierende (outgoing).xlsx": "2-A-8",
    "2-A-9 Ordentliche Studierende (incoming).xlsx": "2-A-9",
    "2-B-1 Doktoratsstudierende mit BV zur Universitaet.xlsx": "2-B-1",
    "3-A-1 Ausserordentliche Studienabschluesse.xlsx": "3-A-1",
    "3-A-2 Studienabschluesse in der Toleranzstudiendauer.xlsx": "3-A-2",
    "3-A-3 Studienabschluesse mit studienbezogenem Auslandsaufenthalt.xlsx": "3-A-3",
}

# Spezial-Kennzahlen mit eigener Konvertierungslogik
SPECIAL_KENNZAHLEN = {"3-A-1", "3-A-3"}


# ============================================================
# HILFSFUNKTIONEN
# ============================================================

def extract_year_from_header(header_text: str) -> Optional[int]:
    """Extrahiert Jahr aus verschiedenen Header-Formaten."""
    if not header_text:
        return None

    text = str(header_text)

    # WS2024 Format
    match = re.search(r'WS(\d{4})', text)
    if match:
        return int(match.group(1))

    # Wintersemester 2024 oder Studienjahr 2024 Format
    match = re.search(r'(?:Wintersemester|Studienjahr)\s*(\d{4})', text)
    if match:
        return int(match.group(1))

    # Nur Jahr
    match = re.search(r'\b(20\d{2})\b', text)
    if match:
        return int(match.group(1))

    return None


def find_header_row(ws, max_rows: int = 30) -> Tuple[Optional[int], Dict[int, int]]:
    """
    Findet die Header-Zeile mit Jahreszahlen.
    Returns: (header_row_idx, {year: col_idx})
    """
    for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True), start=1):
        year_columns = {}
        for col_idx, cell in enumerate(row):
            if cell is not None:
                year = extract_year_from_header(cell)
                if year and year in VALID_YEARS:
                    year_columns[year] = col_idx

        if len(year_columns) >= 2:
            return row_idx, year_columns

    return None, {}


def find_codex_column(ws, header_row: int) -> int:
    """Findet die Spalte mit dem Uni-Code (Codex)."""
    row = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]

    for col_idx, cell in enumerate(row):
        if cell and 'Codex' in str(cell):
            return col_idx

    # Fallback: Spalte B (Index 1)
    return 1


def normalize_value(value) -> Optional[float]:
    """Normalisiert einen Wert zu float oder None."""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def validate_data_point(point: dict) -> bool:
    """Validiert einen einzelnen Datenpunkt."""
    return (
        point.get('uniCode') in VALID_UNI_CODES and
        point.get('year') in VALID_YEARS and
        isinstance(point.get('value'), (int, float, type(None)))
    )


# ============================================================
# STANDARD-KONVERTER
# ============================================================

def convert_standard(filepath: Path, output_dir: Path, kennzahl_code: str) -> dict:
    """
    Standard-Konvertierung fuer die meisten Wissensbilanz-Dateien.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Sheet auswaehlen
    ws = wb['Tab'] if 'Tab' in wb.sheetnames else wb.active

    # Header-Zeile finden
    header_row, year_columns = find_header_row(ws)
    if not year_columns:
        wb.close()
        return {"error": "Keine Jahreszahlen im Header gefunden", "file": filepath.name}

    # Codex-Spalte finden
    codex_col = find_codex_column(ws, header_row)

    data_points = []
    unis_found: Set[str] = set()
    years_found = set(year_columns.keys())
    invalid_count = 0

    # Daten extrahieren
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if codex_col >= len(row):
            continue

        codex = row[codex_col]
        if codex is None:
            continue

        uni_code = str(codex).strip()
        if uni_code not in VALID_UNI_CODES:
            continue

        unis_found.add(uni_code)

        for year, col_idx in year_columns.items():
            if col_idx >= len(row):
                continue

            point = {
                "uniCode": uni_code,
                "year": year,
                "value": normalize_value(row[col_idx]),
                "kennzahl": kennzahl_code
            }

            if validate_data_point(point):
                data_points.append(point)
            else:
                invalid_count += 1

    wb.close()

    if not data_points:
        return {"error": "Keine gueltigen Datenpunkte gefunden", "file": filepath.name}

    # JSON speichern
    output_file = output_dir / f"{kennzahl_code}.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data_points, f, ensure_ascii=False, indent=2)

    return {
        "file": filepath.name,
        "kennzahl": kennzahl_code,
        "output": output_file.name,
        "data_points": len(data_points),
        "invalid_points": invalid_count,
        "universities": len(unis_found),
        "years": sorted(years_found),
        "unis": sorted(unis_found)
    }


# ============================================================
# SPEZIAL-KONVERTER
# ============================================================

def convert_3a1(filepath: Path, output_dir: Path) -> dict:
    """
    Spezial-Konverter fuer 3-A-1 (Ausserordentliche Studienabschluesse).
    Nur fuer Donau-Uni Krems (UM).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Tab']

    data_points = []

    for row_idx in range(17, 100):
        col1 = ws.cell(row=row_idx, column=1).value  # Studienjahr
        col3 = ws.cell(row=row_idx, column=3).value  # Uni-Code
        col5 = ws.cell(row=row_idx, column=5).value  # Wert

        year = extract_year_from_header(col1)

        if year and col3 and 'UM' in str(col3).strip():
            point = {
                "uniCode": "UM",
                "year": year,
                "value": normalize_value(col5),
                "kennzahl": "3-A-1"
            }
            if validate_data_point(point):
                data_points.append(point)

    wb.close()

    output_file = output_dir / "3-A-1.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data_points, f, ensure_ascii=False, indent=2)

    return {
        "file": filepath.name,
        "kennzahl": "3-A-1",
        "output": output_file.name,
        "data_points": len(data_points),
        "universities": 1,
        "years": sorted(set(d['year'] for d in data_points)),
        "unis": ["UM"]
    }


def convert_3a3(filepath: Path, output_dir: Path) -> dict:
    """
    Spezial-Konverter fuer 3-A-3 (Studienabschluesse mit Auslandsaufenthalt).
    Andere Struktur mit Buchstaben-Codes.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Tab']

    # Jahr-Spalten finden (Zeile 11)
    year_columns = {}
    for col in range(1, 20):
        cell = ws.cell(row=11, column=col).value
        year = extract_year_from_header(cell)
        if year:
            year_columns[year] = col + 2  # Gesamt-Spalte ist 2 weiter

    data_points = []
    current_uni_code = None
    unis_found: Set[str] = set()

    for row_idx in range(13, 300):
        uni_name = ws.cell(row=row_idx, column=1).value
        letter = ws.cell(row=row_idx, column=2).value
        category = ws.cell(row=row_idx, column=3).value

        # Neue Uni erkennen
        if uni_name and letter:
            letter_str = str(letter).strip()
            if letter_str in LETTER_TO_UNI_CODE:
                current_uni_code = LETTER_TO_UNI_CODE[letter_str]

        # Nur "Insgesamt"-Zeilen
        if current_uni_code and category and "Insgesamt" in str(category):
            for year, col_idx in year_columns.items():
                point = {
                    "uniCode": current_uni_code,
                    "year": year,
                    "value": normalize_value(ws.cell(row=row_idx, column=col_idx).value),
                    "kennzahl": "3-A-3"
                }
                if validate_data_point(point):
                    data_points.append(point)
                    unis_found.add(current_uni_code)

    wb.close()

    output_file = output_dir / "3-A-3.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data_points, f, ensure_ascii=False, indent=2)

    return {
        "file": filepath.name,
        "kennzahl": "3-A-3",
        "output": output_file.name,
        "data_points": len(data_points),
        "universities": len(unis_found),
        "years": sorted(set(d['year'] for d in data_points)),
        "unis": sorted(unis_found)
    }


# ============================================================
# ANALYSE UND VALIDIERUNG
# ============================================================

def analyze_excel(filepath: Path) -> dict:
    """Analysiert die Struktur einer Excel-Datei."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    result = {"filename": filepath.name, "sheets": []}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row, year_columns = find_header_row(ws)

        sample_rows = []
        for row_idx, row in enumerate(ws.iter_rows(max_row=20, max_col=10, values_only=True)):
            sample_rows.append([str(cell)[:30] if cell is not None else "" for cell in row])

        result["sheets"].append({
            "name": sheet_name,
            "header_row": header_row,
            "years": sorted(year_columns.keys()) if year_columns else [],
            "sample_rows": sample_rows[:15]
        })

    wb.close()
    return result


def validate_output(output_dir: Path) -> dict:
    """Validiert alle generierten JSON-Dateien."""
    results = {"valid": [], "invalid": [], "warnings": []}

    for json_file in output_dir.glob("*.json"):
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)

            if not isinstance(data, list):
                results["invalid"].append(f"{json_file.name}: Kein Array")
                continue

            invalid_points = []
            for i, point in enumerate(data):
                if not validate_data_point(point):
                    invalid_points.append(i)

            if invalid_points:
                results["warnings"].append(
                    f"{json_file.name}: {len(invalid_points)} ungueltige Punkte"
                )
            else:
                results["valid"].append(f"{json_file.name}: {len(data)} Punkte OK")

        except json.JSONDecodeError as e:
            results["invalid"].append(f"{json_file.name}: JSON-Fehler: {e}")
        except Exception as e:
            results["invalid"].append(f"{json_file.name}: Fehler: {e}")

    return results


# ============================================================
# HAUPTFUNKTION
# ============================================================

def convert_file(filepath: Path, output_dir: Path, kennzahl_code: str) -> dict:
    """Konvertiert eine Datei mit dem passenden Konverter."""
    if kennzahl_code == "3-A-1":
        return convert_3a1(filepath, output_dir)
    elif kennzahl_code == "3-A-3":
        return convert_3a3(filepath, output_dir)
    else:
        return convert_standard(filepath, output_dir, kennzahl_code)


def main():
    parser = argparse.ArgumentParser(
        description="Wissensbilanz Excel zu JSON Konverter",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Beispiele:
    python convert.py                    Alle Dateien konvertieren
    python convert.py --file 1-A-1       Nur Kennzahl 1-A-1
    python convert.py --analyze          Strukturen analysieren
    python convert.py --validate         Output validieren
        """
    )
    parser.add_argument("--file", help="Nur diese Kennzahl konvertieren (z.B. 1-A-1)")
    parser.add_argument("--analyze", action="store_true", help="Nur Struktur analysieren")
    parser.add_argument("--validate", action="store_true", help="Output-Dateien validieren")
    parser.add_argument("--verbose", "-v", action="store_true", help="Ausfuehrliche Ausgabe")
    args = parser.parse_args()

    # Pfade
    script_dir = Path(__file__).parent
    project_root = script_dir.parent
    data_dir = project_root / "data"
    output_dir = project_root / "docs" / "data" / "json"
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"[Converter] Wissensbilanz Konverter v2.0")
    print(f"[Converter] Data:   {data_dir}")
    print(f"[Converter] Output: {output_dir}")
    print()

    # Validierung
    if args.validate:
        print("[Converter] Validiere Output...")
        results = validate_output(output_dir)

        for msg in results["valid"]:
            print(f"  [OK] {msg}")
        for msg in results["warnings"]:
            print(f"  [WARN] {msg}")
        for msg in results["invalid"]:
            print(f"  [FAIL] {msg}")

        print(f"\nErgebnis: {len(results['valid'])} OK, "
              f"{len(results['warnings'])} Warnungen, {len(results['invalid'])} Fehler")
        return

    # Dateien sammeln
    all_files = list(data_dir.glob("*.xlsx"))
    if args.file:
        # Kennzahl-Code oder Dateiname
        files = [f for f in all_files if args.file in f.name or
                 FILE_TO_KENNZAHL.get(f.name, "") == args.file]
    else:
        files = [f for f in all_files if f.name in FILE_TO_KENNZAHL]

    if not files:
        print("[Converter] Keine passenden Dateien gefunden")
        print("[Converter] Verfuegbare Dateien:")
        for f in all_files:
            if f.name.startswith(("1-", "2-", "3-")):
                print(f"  - {f.name}")
        return

    print(f"[Converter] {len(files)} Dateien zu verarbeiten")

    # Analyse-Modus
    if args.analyze:
        for filepath in files:
            print(f"\n=== {filepath.name} ===")
            analysis = analyze_excel(filepath)
            for sheet in analysis["sheets"]:
                print(f"Sheet: {sheet['name']}")
                print(f"Header Row: {sheet['header_row']}")
                print(f"Jahre: {sheet['years']}")
                if args.verbose:
                    for i, row in enumerate(sheet["sample_rows"][:8]):
                        non_empty = [c for c in row if c][:4]
                        print(f"  {i+1}: {non_empty}")
        return

    # Konvertierung
    results = []
    for filepath in files:
        kennzahl_code = FILE_TO_KENNZAHL.get(filepath.name)
        if not kennzahl_code:
            print(f"[Converter] Ueberspringe: {filepath.name} (kein Mapping)")
            continue

        print(f"[Converter] {filepath.name} -> {kennzahl_code}.json")

        try:
            result = convert_file(filepath, output_dir, kennzahl_code)
            results.append(result)

            if "error" in result:
                print(f"  [FAIL] {result['error']}")
            else:
                invalid = result.get('invalid_points', 0)
                invalid_msg = f", {invalid} ungueltig" if invalid else ""
                print(f"  [OK] {result['data_points']} Punkte, "
                      f"{result['universities']} Unis, Jahre: {result['years']}{invalid_msg}")

        except Exception as e:
            print(f"  [FAIL] {e}")
            results.append({"file": filepath.name, "error": str(e)})

    # Zusammenfassung
    successful = [r for r in results if "error" not in r]
    total_points = sum(r.get('data_points', 0) for r in successful)

    print(f"\n[Converter] Fertig: {len(successful)}/{len(results)} Dateien konvertiert")
    print(f"[Converter] Gesamt: {total_points} Datenpunkte")
    print(f"[Converter] Datum: {datetime.now().strftime('%Y-%m-%d %H:%M')}")


if __name__ == "__main__":
    main()
