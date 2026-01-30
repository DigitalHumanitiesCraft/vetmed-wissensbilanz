#!/usr/bin/env python3
"""
Spezial-Konverter für 3-A-1 (Außerordentliche Studienabschlüsse).
Diese Kennzahl existiert nur für die Donau-Uni Krems (UM).
"""

import json
from pathlib import Path
import openpyxl
import re


def extract_year(header_text):
    """Extrahiert Jahr aus 'Studienjahr 2023/24'."""
    if header_text and 'Studienjahr' in str(header_text):
        match = re.search(r'Studienjahr\s*(\d{4})', str(header_text))
        if match:
            return int(match.group(1))
    return None


def convert_3a1():
    script_dir = Path(__file__).parent
    project_root = script_dir.parent

    filepath = project_root / "data" / "3-A-1 Außerordentliche Studienabschlüsse.xlsx"
    output_dir = project_root / "docs" / "data" / "json"
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Tab']

    data_points = []

    # Suche nach Zeilen mit "Studienjahr XXXX/XX" und "UM" Code
    for row_idx in range(17, 100):
        col1 = ws.cell(row=row_idx, column=1).value  # Studienjahr
        col3 = ws.cell(row=row_idx, column=3).value  # Uni-Code
        col5 = ws.cell(row=row_idx, column=5).value  # Wert

        year = extract_year(col1)

        if year and col3 and 'UM' in str(col3).strip():
            try:
                value = float(col5) if col5 is not None else None
            except (ValueError, TypeError):
                value = None

            data_points.append({
                "uniCode": "UM",
                "year": year,
                "value": value,
                "kennzahl": "3-A-1"
            })

    wb.close()

    # JSON speichern
    output_file = output_dir / "3-A-1.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data_points, f, ensure_ascii=False, indent=2)

    print(f"Konvertiert: {len(data_points)} Datenpunkte")
    print(f"Gespeichert: {output_file}")

    for dp in data_points:
        print(f"  {dp['year']}: {dp['value']}")


if __name__ == "__main__":
    convert_3a1()
