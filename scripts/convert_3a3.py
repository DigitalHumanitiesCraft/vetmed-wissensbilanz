#!/usr/bin/env python3
"""
Spezial-Konverter für 3-A-3 (Studienabschlüsse mit Auslandsaufenthalt).
Diese Datei hat eine andere Struktur als die Standard-Wissensbilanz-Dateien.
"""

import json
from pathlib import Path
import openpyxl

# Mapping Buchstabe -> Uni-Code
LETTER_TO_CODE = {
    "A": "UA",  # Universität Wien
    "B": "UB",  # Universität Graz
    "C": "UC",  # Universität Innsbruck
    "D": "UD",  # Universität Salzburg
    "E": "UE",  # TU Wien
    "F": "UF",  # TU Graz
    "G": "UG",  # Montanuniversität Leoben
    "H": "UH",  # BOKU
    "I": "UI",  # VetMed
    "J": "UJ",  # WU Wien
    "K": "UK",  # Universität Linz
    "L": "UL",  # Universität Klagenfurt
    "M": "UM",  # Donau-Uni Krems
    "N": "UN",  # Med Uni Wien
    "O": "UO",  # Med Uni Graz
    "Q": "UQ",  # Med Uni Innsbruck
    "R": "UR",  # Akademie der bildenden Künste
    "S": "US",  # Angewandte
    "T": "UT",  # MDW
    "U": "UU",  # Mozarteum
    "V": "UV",  # KUG Graz
    "W": "UW",  # Kunst Uni Linz
}

def extract_year(header_text):
    """Extrahiert Jahr aus 'Studienjahr 2022/23'."""
    if header_text and 'Studienjahr' in str(header_text):
        import re
        match = re.search(r'Studienjahr\s*(\d{4})', str(header_text))
        if match:
            return int(match.group(1))
    return None


def convert_3a3():
    script_dir = Path(__file__).parent
    project_root = script_dir.parent

    filepath = project_root / "data" / "3-A-3 Studienabschlüsse mit studienbezogenem Auslandsaufenthalt.xlsx"
    output_dir = project_root / "docs" / "data" / "json"
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Tab']

    # Header-Zeile 11 hat die Studienjahre
    # Spalte D (3) = 2022/23, Spalte H (7) = 2021/22, etc.
    year_columns = {}
    for col in range(1, 20):
        cell = ws.cell(row=11, column=col).value
        year = extract_year(cell)
        if year:
            # Finde "Gesamt"-Spalte für dieses Jahr (offset +2)
            year_columns[year] = col + 2  # Gesamt ist 2 Spalten weiter

    print(f"Jahre gefunden: {year_columns}")

    data_points = []
    current_uni_code = None

    # Daten ab Zeile 13
    for row_idx in range(13, 300):
        uni_name = ws.cell(row=row_idx, column=1).value
        letter = ws.cell(row=row_idx, column=2).value
        category = ws.cell(row=row_idx, column=3).value

        # Neue Uni beginnt wenn Spalte A gefüllt ist
        if uni_name and letter:
            letter = str(letter).strip()
            if letter in LETTER_TO_CODE:
                current_uni_code = LETTER_TO_CODE[letter]

        # Nur "Insgesamt"-Zeilen extrahieren (Gesamt mit Auslandsaufenthalt)
        if current_uni_code and category and "Insgesamt" in str(category):
            for year, col_idx in year_columns.items():
                value = ws.cell(row=row_idx, column=col_idx).value

                if value is not None:
                    try:
                        value = float(value)
                    except (ValueError, TypeError):
                        value = None

                data_points.append({
                    "uniCode": current_uni_code,
                    "year": year,
                    "value": value,
                    "kennzahl": "3-A-3"
                })

    wb.close()

    # JSON speichern
    output_file = output_dir / "3-A-3.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data_points, f, ensure_ascii=False, indent=2)

    print(f"Konvertiert: {len(data_points)} Datenpunkte")
    print(f"Gespeichert: {output_file}")

    # Unis zeigen
    unis = set(d['uniCode'] for d in data_points)
    print(f"Unis: {sorted(unis)}")


if __name__ == "__main__":
    convert_3a3()
