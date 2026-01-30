#!/usr/bin/env python3
"""
Excel zu JSON Konverter für Wissensbilanz-Daten.

Konvertiert Excel-Dateien aus data/ nach docs/data/json/
Format: Array von {uniCode, year, value, kennzahl}

Verwendung:
    python scripts/convert_excel_to_json.py
    python scripts/convert_excel_to_json.py --file "1-A-1 Personal - Köpfe.xlsx"
    python scripts/convert_excel_to_json.py --analyze  # Nur Struktur analysieren
"""

import os
import sys
import json
import re
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("[Converter] Fehler: openpyxl nicht installiert")
    print("[Converter] Installation: pip install openpyxl")
    sys.exit(1)


# Mapping Dateiname -> Kennzahl-Code
FILE_TO_KENNZAHL = {
    "1-A-1 Personal - Köpfe.xlsx": "1-A-1",
    "1-A-1 Personal - VZÄ.xlsx": "1-A-1-VZA",
    "1-A-2 Berufungen an die Universität.xlsx": "1-A-2",
    "1-A-3 Frauenquote in Kollegialorganen.xlsx": "1-A-3",
    "1-A-4 Gender pay gap.xlsx": "1-A-4",
    "1-A-5 Repräsentanz von Frauen in Berufungsverfahren.xlsx": "1-A-5",
    "2-A-1 ProfessorInnen und Äquivalente.xlsx": "2-A-1",
    "2-A-2 Eingerichtete Studien.xlsx": "2-A-2",
    "2-A-3 Studienabschlussquote.xlsx": "2-A-3",
    "2-A-4 Besondere Zulassungsbedingungen.xlsx": "2-A-4",
    "2-A-5 Anzahl Studierenden.xlsx": "2-A-5",
    "2-A-6 Anzahl Prüfungsaktive.xlsx": "2-A-6",
    "2-A-7 Anzahl belegte ordentliche Studien.xlsx": "2-A-7",
    "2-A-8 Ordentliche Studierende (outgoing).xlsx": "2-A-8",
    "2-A-9 Ordentliche Studierende (incoming).xlsx": "2-A-9",
    "2-B-1 Doktoratsstudierende mit BV zur Universität.xlsx": "2-B-1",
    "3-A-1 Außerordentliche Studienabschlüsse.xlsx": "3-A-1",
    "3-A-2 Studienabschlüsse in der Toleranzstudiendauer.xlsx": "3-A-2",
    "3-A-3 Studienabschlüsse mit studienbezogenem Auslandsaufenthalt.xlsx": "3-A-3",
}

# Offizielle Uni-Codes aus der Wissensbilanz
# Verifiziert aus: 1-A-1 Personal - Köpfe.xlsx
VALID_UNI_CODES = {
    "UA": "Universität Wien",
    "UB": "Universität Graz",
    "UC": "Universität Innsbruck",
    "UD": "Universität Salzburg",
    "UE": "Technische Universität Wien",
    "UF": "Technische Universität Graz",
    "UG": "Montanuniversität Leoben",
    "UH": "Universität für Bodenkultur Wien",
    "UI": "Veterinärmedizinische Universität Wien",
    "UJ": "Wirtschaftsuniversität Wien",
    "UK": "Universität Linz",
    "UL": "Universität Klagenfurt",
    "UM": "Universität für Weiterbildung Krems",
    "UN": "Medizinische Universität Wien",
    "UO": "Medizinische Universität Graz",
    "UQ": "Medizinische Universität Innsbruck",
    "UR": "Akademie der bildenden Künste Wien",
    "US": "Universität für angewandte Kunst Wien",
    "UT": "Universität für Musik und darstellende Kunst Wien",
    "UU": "Universität Mozarteum Salzburg",
    "UV": "Universität für Musik und darstellende Kunst Graz",
    "UW": "Universität für künstlerische und industrielle Gestaltung Linz",
}


def extract_year_from_header(header_text: str) -> int:
    """Extrahiert Jahr aus Header wie 'Wintersemester 2024 (Stichtag: 31.12.2024)'."""
    if not header_text:
        return None

    # WS2024 Format
    match = re.search(r'WS(\d{4})', str(header_text))
    if match:
        return int(match.group(1))

    # Wintersemester 2024 Format
    match = re.search(r'(?:Wintersemester|Studienjahr)\s*(\d{4})', str(header_text))
    if match:
        return int(match.group(1))

    # Nur Jahr
    match = re.search(r'\b(20\d{2})\b', str(header_text))
    if match:
        return int(match.group(1))

    return None


def find_header_row(ws, max_rows=30) -> tuple:
    """
    Findet die Header-Zeile mit Jahreszahlen.
    Returns: (header_row_idx, year_columns: {year: col_idx})
    """
    for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True), start=1):
        year_columns = {}
        for col_idx, cell in enumerate(row):
            if cell is not None:
                year = extract_year_from_header(cell)
                if year and 2015 <= year <= 2030:
                    year_columns[year] = col_idx

        if len(year_columns) >= 2:
            return row_idx, year_columns

    return None, {}


def find_codex_column(ws, header_row: int) -> int:
    """Findet die Spalte mit dem Uni-Code (Codex)."""
    row = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]

    for col_idx, cell in enumerate(row):
        if cell and 'Codex' in str(cell):
            return col_idx

    # Fallback: Spalte 1 (B)
    return 1


def convert_file(filepath: Path, output_dir: Path, kennzahl_code: str) -> dict:
    """
    Konvertiert eine Excel-Datei zu JSON.
    Returns: Statistiken über die Konvertierung
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Versuche Sheet "Tab" oder erstes Sheet
    if 'Tab' in wb.sheetnames:
        ws = wb['Tab']
    else:
        ws = wb.active

    # Header-Zeile finden
    header_row, year_columns = find_header_row(ws)

    if not year_columns:
        wb.close()
        return {"error": "Keine Jahreszahlen im Header gefunden", "file": filepath.name}

    # Codex-Spalte finden
    codex_col = find_codex_column(ws, header_row)

    data_points = []
    unis_found = set()
    years_found = set(year_columns.keys())

    # Daten ab Header+1 extrahieren
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        # Uni-Code aus Codex-Spalte
        if codex_col >= len(row):
            continue

        codex = row[codex_col]
        if codex is None:
            continue

        uni_code = str(codex).strip()

        # Nur gültige 2-Buchstaben Uni-Codes
        if uni_code not in VALID_UNI_CODES:
            continue

        unis_found.add(uni_code)

        # Werte für jedes Jahr extrahieren
        for year, col_idx in year_columns.items():
            if col_idx >= len(row):
                continue

            value = row[col_idx]

            # Wert normalisieren
            if value is not None:
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    value = None

            data_points.append({
                "uniCode": uni_code,
                "year": year,
                "value": value,
                "kennzahl": kennzahl_code
            })

    wb.close()

    if not data_points:
        return {"error": "Keine Datenpunkte gefunden", "file": filepath.name}

    # JSON speichern
    output_file = output_dir / f"{kennzahl_code}.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data_points, f, ensure_ascii=False, indent=2)

    return {
        "file": filepath.name,
        "kennzahl": kennzahl_code,
        "output": output_file.name,
        "data_points": len(data_points),
        "universities": len(unis_found),
        "years": sorted(years_found),
        "unis": sorted(unis_found)
    }


def analyze_excel(filepath: Path) -> dict:
    """Analysiert die Struktur einer Excel-Datei."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    result = {
        "filename": filepath.name,
        "sheets": []
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Header-Zeile finden
        header_row, year_columns = find_header_row(ws)

        # Erste Zeilen analysieren
        sample_rows = []
        for row_idx, row in enumerate(ws.iter_rows(max_row=25, max_col=10, values_only=True)):
            if row_idx < 25:
                sample_rows.append([str(cell)[:40] if cell is not None else "" for cell in row])

        sheet_info = {
            "name": sheet_name,
            "header_row": header_row,
            "years": sorted(year_columns.keys()) if year_columns else [],
            "sample_rows": sample_rows
        }
        result["sheets"].append(sheet_info)

    wb.close()
    return result


def main():
    parser = argparse.ArgumentParser(description="Excel zu JSON Konverter")
    parser.add_argument("--file", help="Nur diese Datei konvertieren")
    parser.add_argument("--analyze", action="store_true", help="Nur Struktur analysieren")
    parser.add_argument("--verbose", "-v", action="store_true", help="Mehr Output")
    args = parser.parse_args()

    # Pfade
    script_dir = Path(__file__).parent
    project_root = script_dir.parent
    data_dir = project_root / "data"
    output_dir = project_root / "docs" / "data" / "json"

    # Output-Verzeichnis erstellen
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"[Converter] data_dir: {data_dir}")
    print(f"[Converter] output_dir: {output_dir}")

    # Dateien sammeln
    if args.file:
        files = [data_dir / args.file]
    else:
        files = [f for f in data_dir.glob("*.xlsx") if f.name in FILE_TO_KENNZAHL]

    if not files:
        print("[Converter] Keine passenden Dateien gefunden")
        print("[Converter] Verfügbare Wissensbilanz-Dateien:")
        for f in data_dir.glob("*.xlsx"):
            if f.name.startswith(("1-", "2-", "3-")):
                print(f"  - {f.name}")
        return

    print(f"[Converter] {len(files)} Dateien zu verarbeiten")

    # Analyse-Modus
    if args.analyze:
        for filepath in files:
            if not filepath.exists():
                print(f"[Converter] Datei nicht gefunden: {filepath}")
                continue

            print(f"\n=== {filepath.name} ===")
            analysis = analyze_excel(filepath)
            for sheet in analysis["sheets"]:
                print(f"\nSheet: {sheet['name']}")
                print(f"Header Row: {sheet['header_row']}")
                print(f"Years: {sheet['years']}")
                if args.verbose:
                    print("Sample rows:")
                    for i, row in enumerate(sheet["sample_rows"][:10]):
                        non_empty = [c for c in row if c][:5]
                        print(f"  {i+1}: {non_empty}")
        return

    # Konvertierung
    results = []
    for filepath in files:
        if not filepath.exists():
            print(f"[Converter] Datei nicht gefunden: {filepath}")
            continue

        kennzahl_code = FILE_TO_KENNZAHL.get(filepath.name)
        if not kennzahl_code:
            print(f"[Converter] Überspringe: {filepath.name} (kein Mapping)")
            continue

        print(f"[Converter] Konvertiere: {filepath.name} -> {kennzahl_code}.json")

        try:
            result = convert_file(filepath, output_dir, kennzahl_code)
            results.append(result)

            if "error" in result:
                print(f"  Fehler: {result['error']}")
            else:
                print(f"  {result['data_points']} Datenpunkte, {result['universities']} Unis, Jahre: {result['years']}")
                if args.verbose:
                    print(f"  Unis: {result['unis']}")
        except Exception as e:
            print(f"  Fehler: {e}")
            results.append({"file": filepath.name, "error": str(e)})

    # Zusammenfassung
    successful = [r for r in results if "error" not in r]
    print(f"\n[Converter] Fertig: {len(successful)}/{len(results)} Dateien konvertiert")

    if successful:
        print("\n[Converter] Generierte JSON-Dateien:")
        for r in successful:
            print(f"  - {r['output']}: {r['data_points']} Datenpunkte")


if __name__ == "__main__":
    main()
